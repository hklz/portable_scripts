from openpyxl import Workbook
from openpyxl import load_workbook
import random
from math import sqrt
#----------------------------------------------------------------#
"""质数检测"""
num = 1
if (num != 1):
    for i in range(2, int(sqrt(num))+1):
        if (num % i == 0):
            print("%d不是素数，最小约数是%d" % (num, i))
            break
    else:
        print("%d是质数"%num)
else:
    print("%d不算质数"%num)

#----------------------------------------------------------------#
#封装成函数
def iFprime(num):
    if (num != 1):
        for i in range(2, int(sqrt(num)) + 1):
            if (num % i == 0):
                return ("%d不是素数，最小约数是%d" % (num, i))
                break
        else:
            return ("%d是质数" % num)
    else:
        return ("%d不算质数" % num)


"""新建excel"""
book = Workbook()
sheet = book.active
#创建新的工作簿
new_sheet = book.create_sheet("New_sheet")
#单元格赋值
sheet['A1'] = 1
sheet['A2'] = 2
#单元格赋一大堆值
for row in range(1,10):
    for col in range(1,10):
        sheet.cell(row=row, column=col).value =random.randint(0,1000)
#工作簿保存
book.save("test.xlsx")
#----------------------------------------------------------------#

"""从已有xlsx中读取数据"""
#打开工作簿
open_xlsx = load_workbook("test.xlsx")
#输出工作簿里面工作表的名字
print(open_xlsx.sheetnames)
#打开工作表
open_xlsx_sheet = open_xlsx["Sheet"]
#访问一个格子
open_xlsx_sheet_cell_A1 = open_xlsx_sheet['A1']
open_xlsx_sheet_cell_A2 = open_xlsx_sheet.cell(row=1, column=2)
#访问格子的值
open_xlsx_sheet_cell_A1_value = open_xlsx_sheet_cell_A1.value
open_xlsx_sheet_cell_A2_value = open_xlsx_sheet.cell(row=2, column=1).value
print(open_xlsx_sheet_cell_A1_value)
print(open_xlsx_sheet_cell_A2_value)
#读一大堆值
for row in range(1,10):
    for col in range(1,10):
        print(open_xlsx_sheet.cell(row=row, column=col).value, end="\t")
        print(iFprime(open_xlsx_sheet.cell(row=row, column=col).value))
print()
open_xlsx.close()



